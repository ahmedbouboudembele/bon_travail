import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
import mysql.connector
import hashlib
import io
from openpyxl import Workbook
from openpyxl.styles import Font

# ---------------------------
# Config page
# ---------------------------
st.set_page_config(page_title="Gestion Bons de Travail", layout="wide")
st.title("Gestion des Bons de Travail - MySQL")

# ---------------------------
# MySQL connection
# ---------------------------
def get_connection():
    return mysql.connector.connect(
        host="127.0.0.1",
        user="root",
        password="",
        database="bon_travail_db"
    )

# ---------------------------
# Password hashing
# ---------------------------
def hash_password(pwd: str) -> str:
    return hashlib.sha256(pwd.encode('utf-8')).hexdigest()

# ---------------------------
# Users / Auth functions
# ---------------------------
def get_user(username):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM users WHERE username=%s", (username,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return user

def create_user(username, password, role):
    conn = get_connection()
    cursor = conn.cursor()
    hashed = hash_password(password)
    cursor.execute("INSERT INTO users (username, password_hash, role) VALUES (%s,%s,%s)",
                   (username, hashed, role))
    conn.commit()
    cursor.close()
    conn.close()

# ---------------------------
# Bons de travail CRUD
# ---------------------------
BON_COLUMNS = [
    "code","date","arret_declare_par","poste_de_charge","heure_declaration","machine_arreter",
    "heure_debut_intervention","heure_fin_intervention","technicien","description_probleme",
    "action","pdr_utilisee","observation","resultat","condition_acceptation",
    "dpt_maintenance","dpt_qualite","dpt_production"
]

def read_bons(filters=None):
    conn = get_connection()
    df = pd.read_sql("SELECT * FROM bon_travail_db", conn)
    conn.close()
    if filters:
        for k,v in filters.items():
            if v:
                df = df[df[k]==v]
    return df

def add_bon(row):
    conn = get_connection()
    cursor = conn.cursor()
    cols = ",".join(row.keys())
    vals = tuple(row.values())
    placeholders = ",".join(["%s"]*len(row))
    cursor.execute(f"INSERT INTO bon_travail_db ({cols}) VALUES ({placeholders})", vals)
    conn.commit()
    cursor.close()
    conn.close()

def update_bon(code, row):
    conn = get_connection()
    cursor = conn.cursor()
    set_str = ", ".join([f"{k}=%s" for k in row.keys()])
    vals = tuple(row.values()) + (code,)
    cursor.execute(f"UPDATE bon_travail_db SET {set_str} WHERE code=%s", vals)
    conn.commit()
    cursor.close()
    conn.close()

def delete_bon(code):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM bon_travail_db WHERE code=%s", (code,))
    conn.commit()
    cursor.close()
    conn.close()

# ---------------------------
# PDR CRUD
# ---------------------------
PDR_COLUMNS = ["code","remplacement","nom_composant","quantite"]

def read_pdr():
    conn = get_connection()
    df = pd.read_sql("SELECT * FROM liste_pdr", conn)
    conn.close()
    return df

def upsert_pdr(rec):
    conn = get_connection()
    cursor = conn.cursor()
    code = rec["code"]
    cursor.execute("SELECT * FROM liste_pdr WHERE code=%s", (code,))
    if cursor.fetchone():
        cursor.execute("UPDATE liste_pdr SET remplacement=%s, nom_composant=%s, quantite=%s WHERE code=%s",
                       (rec["remplacement"], rec["nom_composant"], int(rec["quantite"]), code))
    else:
        cursor.execute("INSERT INTO liste_pdr (code, remplacement, nom_composant, quantite) VALUES (%s,%s,%s,%s)",
                       (code, rec["remplacement"], rec["nom_composant"], int(rec["quantite"])))
    conn.commit()
    cursor.close()
    conn.close()

def delete_pdr(code):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM liste_pdr WHERE code=%s", (code,))
    conn.commit()
    cursor.close()
    conn.close()

# ---------------------------
# Excel export
# ---------------------------
def export_excel(bons_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bons de travail"
    for col_idx, h in enumerate(BON_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx).value = h
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    for row_idx, (_, row) in enumerate(bons_df.iterrows(), start=2):
        for col_idx, h in enumerate(BON_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx).value = row[h]
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------------------
# Session state
# ---------------------------
if "user" not in st.session_state:
    st.session_state.user = None
if "role" not in st.session_state:
    st.session_state.role = None
if "manager_verified" not in st.session_state:
    st.session_state.manager_verified = False

# ---------------------------
# Sidebar: login & user management
# ---------------------------
st.sidebar.title("Connexion")
if st.session_state.user:
    st.sidebar.success(f"Connecté: {st.session_state.user} ({st.session_state.role})")
    if st.sidebar.button("Se déconnecter"):
        st.session_state.user = None
        st.session_state.role = None
        st.rerun()
else:
    login_user = st.sidebar.text_input("Nom d'utilisateur")
    login_pwd = st.sidebar.text_input("Mot de passe", type="password")
    if st.sidebar.button("Se connecter"):
        u = get_user(login_user)
        if not u or u["password_hash"] != hash_password(login_pwd):
            st.sidebar.error("Identifiants invalides")
        else:
            st.session_state.user = u["username"]
            st.session_state.role = u["role"]
            st.sidebar.success(f"Bienvenue {u['username']} ({u['role']})")
            st.rerun()

# ---------------------------
# Manager: création utilisateur
# ---------------------------
st.sidebar.markdown("---")
st.sidebar.subheader("Créer un compte (manager)")
if not st.session_state.manager_verified:
    mgr_name = st.sidebar.text_input("Manager", key="mgr_name")
    mgr_pwd = st.sidebar.text_input("Password manager", type="password", key="mgr_pwd")
    if st.sidebar.button("Vérifier manager"):
        mgr = get_user(mgr_name)
        if not mgr or mgr["password_hash"] != hash_password(mgr_pwd) or mgr["role"] != "manager":
            st.sidebar.error("Vérification échouée")
        else:
            st.sidebar.success("Manager vérifié")
            st.session_state.manager_verified = True
            st.rerun()
else:
    new_user = st.sidebar.text_input("Nouvel utilisateur")
    new_pwd = st.sidebar.text_input("Mot de passe", type="password")
    new_role = st.sidebar.selectbox("Rôle", ["production","maintenance","qualite","manager"])
    if st.sidebar.button("Créer utilisateur"):
        create_user(new_user.strip(), new_pwd, new_role)
        st.sidebar.success("Utilisateur créé")
        st.session_state.manager_verified = False

# ---------------------------
# Pages & permissions
# ---------------------------
menu = st.sidebar.radio("Pages", ["Dashboard","Bons","PDR","Export Excel"])

def allowed(page):
    role = st.session_state.role
    if role=="manager": return True
    if role=="production" and page=="Bons": return True
    if role=="maintenance" and page=="Bons": return True
    if role=="qualite" and page=="Bons": return True
    return False

# ---------------------------
# Page: Dashboard
# ---------------------------
def page_dashboard():
    st.header("Dashboard")
    bons_df = read_bons()
    if bons_df.empty:
        st.info("Aucun bon enregistré")
        return
    st.subheader("Bons de travail récents")
    st.dataframe(bons_df.sort_values("date", ascending=False).head(20))
    
    st.subheader("Pareto des problèmes")
    top_action = bons_df['description_probleme'].value_counts().head(10)
    fig, ax = plt.subplots()
    top_action.plot(kind='bar', ax=ax)
    ax.set_ylabel("Nombre")
    st.pyplot(fig)

# ---------------------------
# Page: Bons
# ---------------------------
def page_bons():
    if not allowed("Bons"):
        st.warning("Pas de permission")
        return
    st.header("Gestion des Bons de Travail")
    
    bons_df = read_bons()
    
    st.subheader("Filtres")
    col1,col2,col3 = st.columns(3)
    with col1:
        filt_tech = st.selectbox("Technicien", [""] + bons_df['technicien'].dropna().unique().tolist())
    with col2:
        filt_date = st.date_input("Date", value=None)
    with col3:
        filt_dept = st.selectbox("Département", [""] + ["dpt_maintenance","dpt_qualite","dpt_production"])
    
    filters = {}
    if filt_tech: filters['technicien'] = filt_tech
    if filt_date: filters['date'] = filt_date
    if filt_dept: filters[filt_dept] = "Oui"
    
    bons_df_filtered = read_bons(filters)
    st.dataframe(bons_df_filtered)
    
    st.subheader("Ajouter / Modifier un bon")
    with st.form("form_bon"):
        code = st.text_input("Code")
        date_b = st.date_input("Date")
        technicien = st.text_input("Technicien")
        desc = st.text_area("Description problème")
        action = st.text_area("Action")
        pdr = st.text_input("PDR utilisée")
        res = st.text_input("Résultat")
        if st.form_submit_button("Ajouter/Mettre à jour"):
            row = {
                "code": code,
                "date": date_b,
                "arret_declare_par": st.session_state.user,
                "poste_de_charge": "",
                "heure_declaration": "",
                "machine_arreter": "",
                "heure_debut_intervention": "",
                "heure_fin_intervention": "",
                "technicien": technicien,
                "description_probleme": desc,
                "action": action,
                "pdr_utilisee": pdr,
                "observation": "",
                "resultat": res,
                "condition_acceptation": "",
                "dpt_maintenance": "Oui" if st.session_state.role=="maintenance" else "",
                "dpt_qualite": "Oui" if st.session_state.role=="qualite" else "",
                "dpt_production": "Oui" if st.session_state.role=="production" else ""
            }
            existing = read_bons({"code": code})
            if not existing.empty:
                update_bon(code, row)
                st.success("Bon mis à jour")
            else:
                add_bon(row)
                st.success("Bon ajouté")
            st.rerun()

# ---------------------------
# Page: PDR
# ---------------------------
def page_pdr():
    st.header("Liste PDR")
    pdr_df = read_pdr()
    st.dataframe(pdr_df)
    
    st.subheader("Ajouter / Modifier PDR")
    with st.form("form_pdr"):
        code = st.text_input("Code PDR")
        remplacement = st.text_input("Remplacement")
        nom_composant = st.text_input("Nom composant")
        quantite = st.number_input("Quantité", min_value=0, step=1)
        if st.form_submit_button("Ajouter / Modifier"):
            rec = {"code": code,"remplacement": remplacement,"nom_composant": nom_composant,"quantite": quantite}
            upsert_pdr(rec)
            st.success("PDR enregistrée")
            st.rerun()

# ---------------------------
# Page: Export Excel
# ---------------------------
def page_export():
    st.header("Export Excel")
    bons_df = read_bons()
    if bons_df.empty:
        st.info("Aucun bon")
        return
    excel_bytes = export_excel(bons_df)
    st.download_button("Télécharger Excel", excel_bytes, file_name="bons_travail.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------
# Router
# ---------------------------
if menu=="Dashboard":
    page_dashboard()
elif menu=="Bons":
    page_bons()
elif menu=="PDR":
    page_pdr()
elif menu=="Export Excel":
    page_export()
